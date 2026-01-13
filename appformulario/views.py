from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, permission_required
from django.db.models import ProtectedError
from django.http import HttpResponse
from django.contrib import messages
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.core.paginator import Paginator
from django.db.models import Q

from AppMunicipio.models import Municipio
from AppPuestoVotacion.models import PuestoVotacion
from appcorregimientos.models import Corregimiento

from .forms import VotanteForm
from .models import Votante
from django.http import JsonResponse
from appmesa.models import Mesa
from io import BytesIO
import pandas as pd


@login_required
@permission_required('appformulario.view_votante', raise_exception=True)
def lista_votantes(request):
    query = request.GET.get('q')
    status_filter = request.GET.get('status')
    rol_filter = request.GET.get('rol')
    lider_filter = request.GET.get('lider') 
    
    # Es importante ordenar los resultados para que la paginación sea consistente
    votantes_list = Votante.objects.all().order_by('-id') # O por 'apellido'

    # Aplicar el filtro de estado si existe
    if status_filter:
        votantes_list = votantes_list.filter(status=status_filter)

    # Aplicar el filtro de rol si existe
    if rol_filter:
        votantes_list = votantes_list.filter(rol=rol_filter)

    if lider_filter:
        votantes_list = votantes_list.filter(lider_asignado_id=lider_filter)
    
    # Aplicar el filtro de búsqueda general (q)
    if query:
        votantes_list = votantes_list.filter(
            Q(nombre__icontains=query) |
            Q(apellido__icontains=query) |
            Q(cedula__icontains=query) |
            Q(puesto_votacion__nombre_lugar__icontains=query)
        )


    lideres = Votante.objects.filter(
        rol='LIDER_VOTANTE',
        votantes_asignados__isnull=False
    ).distinct()
    
    # --- LÓGICA DE PAGINACIÓN ---
    paginator = Paginator(votantes_list, 10) # Muestra 10 votantes por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'votantes/lista.html', {
        'votantes': page_obj, # Ahora pasamos el objeto paginado
        'query': query,
        'status_filter': status_filter,
        'rol_filter': rol_filter,
        'lider_filter': lider_filter,
        'lideres': lideres,
    })

@login_required
@permission_required('appformulario.add_votante', raise_exception=True)
def crear_votante(request):
    if request.method == 'POST':
        form = VotanteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ El votante fue registrado exitosamente.")
            return redirect('lista_votantes')
    else:
        form = VotanteForm()

    return render(request, 'votantes/form.html', {'form': form, 'titulo': 'Registrar Votante'})


@login_required
@permission_required('appformulario.change_votante', raise_exception=True)
def editar_votante(request, pk):
    vot = get_object_or_404(Votante, pk=pk)
    if request.method == 'POST':
        form = VotanteForm(request.POST, instance=vot)
        if form.is_valid():
            form.save()
            messages.success(request, "✅ Los datos del votante fueron actualizados correctamente.")
            return redirect('lista_votantes')
    else:
        form = VotanteForm(instance=vot)

    return render(request, 'votantes/form.html', {'form': form, 'titulo': 'Editar Votante'})


@login_required
@permission_required('appformulario.delete_votante', raise_exception=True)
def eliminar_votante(request, pk):
    vot = get_object_or_404(Votante, pk=pk)
    try:
        vot.delete()
        messages.success(request, "✅ Votante eliminado.")
    except ProtectedError:
        messages.error(request, "❌ No se puede eliminar este votante.")
    return redirect('lista_votantes')


@login_required
@permission_required('appformulario.change_status_votante', raise_exception=True)
def cambiar_estado_votante(request, pk):
    vot = get_object_or_404(Votante, pk=pk)
    
    if request.method == 'POST':
        # --- CASO: INACTIVACIÓN (Viene del Modal) ---
        if vot.status == 'ACTIVE':
            motivo = request.POST.get('motivo')
            
            if not motivo:
                messages.error(request, "❌ Debe proporcionar un motivo para inactivar al votante.")
            else:
                vot.status = 'INACTIVE'
                vot.motivo_inactivacion = motivo  # Guardamos la razón
                vot.save()
                messages.warning(request, f"⚠️ Votante {vot.nombre} marcado como INACTIVO por: {motivo}")
    
    else:
        # --- CASO: ACTIVACIÓN (Viene del enlace simple GET) ---
        if vot.status == 'INACTIVE':
            vot.status = 'ACTIVE'
            vot.motivo_inactivacion = None  # Limpiamos el motivo al reactivar
            vot.save()
            messages.success(request, f"✅ Votante {vot.nombre} {vot.apellido} ha sido REACTIVADO.")
        else:
            # Si intentan acceder por URL siendo ya activo, no hacemos nada o redirigimos
            pass

    # Redirige a la página anterior o a la lista por defecto
    return redirect(request.META.get('HTTP_REFERER', 'lista_votantes'))

@login_required
def obtener_mesas_por_puesto(request):
    puesto_id = request.GET.get("puesto_id")
    mesas = Mesa.objects.filter(puesto_votacion_id=puesto_id, status="ACTIVE")

    data = [
        {"id": m.id, "numero": m.numero}
        for m in mesas
    ]
    return JsonResponse({"mesas": data})


@login_required
def obtener_corregimientos_por_municipio(request):
    municipio_id = request.GET.get("municipio_id")
    data = []

    if municipio_id:
        corregimientos = Corregimiento.objects.filter(
            municipio_id=municipio_id,
            status="ACTIVE"
        ).order_by("nombre")

        for c in corregimientos:
            data.append({
                "id": c.id,
                "nombre": c.nombre
            })

    return JsonResponse({"corregimientos": data})

@login_required
def obtener_municipios_por_departamento(request):
    departamento_id = request.GET.get("departamento_id")
    data = []

    if departamento_id:
        municipios = Municipio.objects.filter(departamento_id=departamento_id).order_by("nombre")
        for m in municipios:
            data.append({"id": m.id, "nombre": m.nombre})

    return JsonResponse({"municipios": data})

# views.py

@login_required
def obtener_puestos_por_ubicacion(request):
    municipio_id = request.GET.get("municipio_id")
    corregimiento_id = request.GET.get("corregimiento_id")
    
    puestos = PuestoVotacion.objects.filter(status="ACTIVE")

    if corregimiento_id:
        puestos = puestos.filter(corregimiento_id=corregimiento_id)
    elif municipio_id:
        puestos = puestos.filter(municipio_id=municipio_id)

    data = [
        {
            "id": p.id, 
            "nombre": p.nombre_lugar,
            "direccion": p.direccion,
            "municipio": p.municipio.nombre,
            "departamento": p.municipio.departamento.nombre
        } for p in puestos
    ]
    return JsonResponse({"puestos": data})


# -----------------------------------------------------------------------
# Función para exportar a Excel
@login_required
@permission_required('appformulario.export_excel_votante', raise_exception=True)
def exportar_votantes_excel(request):

    # Función para sustituir vacíos por '-'
    def safe(value):
        return value if value not in [None, '', ' '] else '-'

    # Obtener solo votantes activos
    votantes = Votante.objects.filter(status='ACTIVE').select_related(
        'departamento_residencia',
        'municipio_residencia',
        'corregimiento_residencia',
        'puesto_votacion',
        'puesto_votacion__departamento',
        'puesto_votacion__municipio',
        'puesto_votacion__corregimiento',
        'mesa',
        'lider_asignado'
    )

    data = []
    for v in votantes:

        puesto = v.puesto_votacion

        nombre_puesto = puesto.nombre_lugar if puesto else None
        direccion_puesto = puesto.direccion if puesto else None
        municipio_puesto = puesto.municipio.nombre if puesto and puesto.municipio else None
        departamento_puesto = (
            puesto.municipio.departamento.nombre
            if puesto and puesto.municipio and puesto.municipio.departamento
            else None
        )

        data.append({
            # ----- DATOS DEL VOTANTE -----
            'Rol': safe(v.get_rol_display()),
            'Nombres': safe(v.nombre),
            'Apellidos': safe(v.apellido),
            'Cédula': safe(v.cedula),

            # ----- RESIDENCIA -----
            'Departamento residencia': safe(
                v.departamento_residencia.nombre if v.departamento_residencia else None
            ),
            'Municipio residencia': safe(
                v.municipio_residencia.nombre if v.municipio_residencia else None
            ),
            'Corregimiento residencia': safe(
                v.corregimiento_residencia.nombre if v.corregimiento_residencia else None
            ),
            'Dirección residencia': safe(v.direccion_residencia),
            'Barrio residencia': safe(v.barrio_residencia),
            'Teléfono': safe(v.telefono),

            # ----- PUESTO DE VOTACIÓN -----
            'Puesto de votación': safe(
                v.puesto_votacion.nombre_lugar if v.puesto_votacion else None
            ),
            'Dirección puesto': safe(
                v.puesto_votacion.direccion if v.puesto_votacion else None
            ),
            'Departamento puesto': safe(
                v.puesto_votacion.departamento.nombre
                if v.puesto_votacion and v.puesto_votacion.departamento else None
            ),
            'Municipio puesto': safe(
                v.puesto_votacion.municipio.nombre
                if v.puesto_votacion and v.puesto_votacion.municipio else None
            ),
            'Corregimiento puesto': safe(
                v.puesto_votacion.corregimiento.nombre
                if v.puesto_votacion and v.puesto_votacion.corregimiento else None
            ),

            # ----- MESA -----
            'Mesa': safe(v.mesa.numero if v.mesa else None),

            # ----- LÍDER -----
            'Líder asignado': safe(
                f"{v.lider_asignado.nombre} {v.lider_asignado.apellido}"
                if v.lider_asignado else None
            ),
        })

    # Crear DataFrame
    df = pd.DataFrame(data)
    buffer = BytesIO()

    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Votantes')

    buffer.seek(0)
    wb = load_workbook(buffer)
    ws = wb.active

    # -----------------------------------------------
    # ESTILOS
    # -----------------------------------------------
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    ws.freeze_panes = 'A2'

    # ----- ENCABEZADOS -----
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws.auto_filter.ref = ws.dimensions

    # ----- FILAS -----
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = zebra_fill

        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )

    # ----- AUTO-ANCHO -----
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter

        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))

        max_len = min(max_len, 40)
        ws.column_dimensions[col_letter].width = max_len + 3

    # Exportar archivo
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=Listado_Votantes.xlsx'
    return response
